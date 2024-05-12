from decimal import Decimal
import time
from django.forms import formset_factory
from django.shortcuts import render, redirect, get_object_or_404, reverse
from django.http import HttpResponse
from django.views import View
from .models import CustomerAccount, LineItem, Invoice,Recovery,Transaction
from .forms import LineItemForm, LineItemFormset, InvoiceForm,RecoveryForm 
from myapp.models import Product, Shop
from django.utils import timezone
# from django.contrib import messages
from django.contrib import messages
from datetime import datetime, timedelta
from django.http import JsonResponse
from django.db import transaction
from django.db.models import F
from django.db.models import Sum
from openpyxl import Workbook


import pdfkit

def manager_approval(request, recovery_id, invoice_id):
    recovery = get_object_or_404(Recovery, id=recovery_id)
    invoice = get_object_or_404(Invoice, id=invoice_id)

    if request.method == 'POST':
        # Check if the manager_approval_checkbox is checked
        if request.POST.get('manager_approval_checkbox'):
            # Check if manager approval is already approved
            if not recovery.approved_by_manager:
                recovery.approved_by_manager = True
                recovery.save()

    # Filter approved recoveries
    approved_recoveries = Recovery.objects.filter(approved_by_manager=True, invoice__customer=invoice.customer)

    # Calculate total received amount for all approved recoveries related to the customer
    total_received_amount = approved_recoveries.aggregate(total_received=Sum('amount_received'))['total_received']
    total_received_amount = total_received_amount or 0

    # Delete existing transaction for this customer, if it exists
    existing_transaction = Transaction.objects.filter(customer=invoice.customer).first()
    if existing_transaction:
        existing_transaction.delete()

    # Create a new Transaction instance
    Transaction.objects.create(
        recovery=recovery,
        customer=invoice.customer,
        manager=invoice.manager,
        salesperson=invoice.salesperson,
        due_date=invoice.due_date,
        routing=invoice.routing,
        total_amount=recovery.total_amount,
        amount=total_received_amount
    )
    # return redirect('invoice:recovery-list')
    # Redirect to the recovery-list URL after processing
    # redirect('invoice:recovery-list')

    return render(request, "invoice/manager_approval.html", {'recovery': recovery, 'invoice': invoice})

   


class InvoiceListView(View):
    def get(self, *args, **kwargs):
        invoices = Invoice.objects.all().order_by("-id")
        context = {
            "invoices":invoices,
        }

        return render(self.request, 'invoice/invoice-list.html', context)
    
    # def post(self, request):        
    #     # import pdb;pdb.set_trace()
    #     invoice_ids = request.POST.getlist("invoice_id")
    #     invoice_ids = list(map(int, invoice_ids))

    #     update_status_for_invoices = int(request.POST['status'])
    #     invoices = Invoice.objects.filter(id__in=invoice_ids)
    #     # import pdb;pdb.set_trace()
    #     if update_status_for_invoices == 0:
    #         invoices.update(status=False)
    #     else:
    #         invoices.update(status=True)

        # return redirect('invoice:invoice-list')

def createInvoice(request):
    """
    Invoice Generator page it will have Functionality to create new invoices, 
    this will be protected view, only admin has the authority to read and make
    changes here.
    """
    heading_message = 'Formset Demo'
    if request.method == 'GET':
        formset = LineItemFormset(request.GET or None)
        form = InvoiceForm(request.GET or None)
    elif request.method == 'POST':
        formset = LineItemFormset(request.POST)
        form = InvoiceForm(request.POST)
        
        if form.is_valid() and formset.is_valid():
            due_date = form.cleaned_data['due_date']
            if due_date == '':
                due_date = datetime.now().date()
            invoice  = Invoice.objects.create(customer=form.data["customer"],
                    customer_email=form.data["customer_email"],
                    billing_address = form.data["billing_address"],
                    # date=form.data["date"],
                    due_date=due_date,
                    salesperson=form.data["salesperson"],
                    manager=form.data["manager"],
                    sale_types=form.data['sale_types'],
                    routing=form.data["routing"],
                    ) 
        if formset.is_valid():
            # import pdb;pdb.set_trace()
            # extract name and other data from each form and save
            total = Decimal(0)  # Initialize total as Decimal
            # invalid_quantity = False  # Flag to track if any quantity is invalid
            for form in formset:
                service = form.cleaned_data.get('service')
                # description = form.cleaned_data.get('description')
                quantity = form.cleaned_data.get('quantity')
                rate = form.cleaned_data.get('rate')
                if service and quantity and rate:
                    amount = Decimal(str(rate)) * Decimal(str(quantity))  # Convert rate and quantity to Decimal objects
                    # amount = float(rate)*float(quantity)
                    total += amount
                    LineItem(customer=invoice,
                            service=service,
                            # description=description,
                            quantity=quantity,
                            rate=rate,
                            amount=amount).save()
                # else:
                #     invalid_quantity = True  # Set flag to True if any quantity is invalid
                #     break  # Exit the loop as soon as an invalid quantity is encountered
                    # print(total)
            invoice.total_amount = total
            if invoice.total_amount==0:
                 messages.error(request, 'Please Enter correct Quantity')
            else:
                invoice.save()
                messages.success(request, 'Invoice created successfully.')
                request.session['show_success_message'] = True  # Set flag in session
                # time.sleep(5)
                return redirect('/invoice/')

        else:
            messages.error(request, 'Invalid Quantity. Please enter correct correct quantity')

    all_messages = messages.get_messages(request)
    context = {
        "title" : "Invoice Generator",
        "formset": formset,
        "form": form,
        'messages':all_messages
    }
    return render(request, 'invoice/invoice-create.html', context)


def editInvoice(request, id):
    """
    Invoice editing page.
    Only admin has the authority to read and make changes here.
    """
    invoice = get_object_or_404(Invoice, pk=id)
    try:
        if request.method == 'POST':
            # Update the attributes of the invoice object
            invoice.customer = request.POST.get('customer')
            invoice.customer_email = request.POST.get('customer_email')
            invoice.billing_address = request.POST.get('billing_address')
            # invoice.date = request.POST.get('date')
            invoice.date = datetime.strptime(request.POST.get('date'), '%Y-%m-%d')
            # invoice.due_date = request.POST.get('due_date')
            invoice.due_date = datetime.strptime(request.POST.get('due_date'), '%Y-%m-%d')
            # invoice.message = request.POST.get('message')
            invoice.salesperson = request.POST.get('salesperson')
            invoice.manager = request.POST.get('manager')
            invoice.routing = request.POST.get('routing')
            invoice.total_amount = request.POST.get('total_amount')
            # invoice.paid_amount = request.POST.get('paid_amount')
            # invoice.balance = request.POST.get('balance')
            # invoice.previous_balance = request.POST.get('previous_balance')
            # Update more fields as needed
            # Save the changes
            invoice.save()
            return redirect('/invoice/')
    except Exception as e:
            messages.error(request, f'An error occurred: {e}')
            # return JsonResponse({'error': str(e)}, status=500)

    # Get all messages and pass them to the template context
    all_messages = messages.get_messages(request)
    context = {
        'invoice': invoice,
        'invoice_id': id,
        'messages':all_messages
    }
    return render(request, 'invoice/invoice-edit.html', context)


def delete_invoice(request, id):
    invoice = get_object_or_404(Invoice, id=id)
    try:
        if request.method == 'POST':
            invoice.delete()
            return redirect('/invoice/')
    except Exception as e:
                messages.error(request, f'An error occurred: {e}')
                return JsonResponse({'error': str(e)}, status=500)
    
    return render(request, 'invoice/delete_invoice_confirm.html', {'invoice': invoice})


def view_PDF(request, id=None):
    invoice = get_object_or_404(Invoice, id=id)
    lineitem = invoice.lineitem_set.all()

    today_date = datetime.now()
    # Extract day, month, and year components from the due_date
    invoice_date = today_date
    day = invoice_date.day
    month = invoice_date.month
    year = invoice_date.year


    context = {
        "company": {
            "name": "Sh.A.Rehman Traders",
            "address" :"Main Bazaar Piplan (Mianwali)",
            "phone": "03006084456",
            # "email": "asad@gmail.com",
        },
        "invoice_id": f"{invoice.id}-{day}-{month}{year}",
        "invoice_total": invoice.total_amount,
        "customer": invoice.customer,
        "customer_email": invoice.customer_email,
        "date": invoice_date,
        "due_date": invoice.due_date,
        "billing_address": invoice.billing_address,
        "message": "---",
        "Sale_Type":invoice.sale_types,
        "lineitem": lineitem,

    }
    return render(request, 'invoice/pdf_template.html', context)

def generate_PDF(request, id):
    # Use False instead of output path to save pdf to a variable
    config = pdfkit.configuration(wkhtmltopdf='C:\\Program Files\\wkhtmltopdf\\bin\wkhtmltopdf.exe')
    pdf = pdfkit.from_url(request.build_absolute_uri(reverse('invoice:invoice-detail', args=[id])), False, configuration=config)
    response = HttpResponse(pdf,content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="invoice.pdf"'

    return response


def view_404(request,  *args, **kwargs):

    return redirect('invoice:invoice-list')


# from urllib.parse import unquote

def get_details_customer(request, shop_name):
    try:
        shop = Shop.objects.get(name=shop_name)
        data = {
            'name': shop.name,
            'address': shop.address,
            'owner_number': shop.owner_number,
            'owner_cnic': shop.owner_cnic,
        }
        return JsonResponse(data)
    except Shop.DoesNotExist:
        return JsonResponse({'error': 'Shop not found'}, status=404)
    
# def get_details_customer(request, shop_name):
#     # Decode the shop_name
#     decoded_shop_name = unquote(shop_name)
    
#     try:
#         shop = Shop.objects.get(name=decoded_shop_name)
#         data = {
#             'name': shop.name,
#             'address': shop.address,
#             'owner_number': shop.owner_number,
#             'owner_cnic': shop.owner_cnic,
#         }
#         return JsonResponse(data)
#     except Shop.DoesNotExist:
#         return JsonResponse({'error': 'Shop not found'}, status=404)

# Showing httpresponse when i hit /invoice/get_details_customer/zulfiqarali in browser  
# def get_details_customer(request, shop_name):
#     try:
#         shop = get_object_or_404(Shop, name=shop_name)
#         data = {
#             'name': shop.name,
#             'address': shop.address,
#             'owner_number': shop.owner_number,
#             'owner_cnic': shop.owner_cnic,
#         }
#         return HttpResponse(json.dumps(data), content_type='application/json')
#     except Http404:
#         return HttpResponse(json.dumps({'error': 'Shop not found'}), status=404, content_type='application/json')

    
def get_product_rate(request, product_name):
    try:
        product = Product.objects.get(product_name=product_name)
        data = {
            'rate': product.price,  # Assuming 'price' is the rate field
        }
        return JsonResponse(data)
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)
    

# def invoice_list_recovery(request):
#     invoices = Invoice.objects.all()
#     return render(request, 'invoice/invoice_list_recovery.html', {'invoices': invoices})




def create_recovery(request, customer_name):
    # Retrieve the invoices for the given customer name
    invoices = Invoice.objects.filter(customer=customer_name)
    
    # Check if any invoices are found for the customer
    if not invoices:
        return JsonResponse({'message': 'No invoices found for the customer'})
    
    # Calculate the total amount of all invoices for the customer
    total_amount = invoices.aggregate(total_sale_amount=Sum('total_amount'))['total_sale_amount']
    
    # Calculate the total amount received for all invoices for the customer
    total_received_amount = Recovery.objects.filter(invoice__customer=customer_name).aggregate(total_received_amount=Sum('amount_received'))['total_received_amount']
    
    # Check if the total amount received equals the total amount
    if total_received_amount == total_amount:
        return JsonResponse({'message': 'Sale and recoveries are equal'})
    
    thank = False
    if request.method == "POST":
        # Retrieve the total amount for the customer again as it might have changed
        total_amount = invoices.aggregate(total_sale_amount=Sum('total_amount'))['total_sale_amount']
        
        # Retrieve the amount received from the form
        amount_received_str = request.POST.get('new_paid_amount', '')
        new_amount_received = Decimal(amount_received_str)

        # Get the last recovery for the customer
        last_recovery = Recovery.objects.filter(customer=customer_name).order_by('-id').first()
        
        # Initialize previous_balance with the balance of the last recovery or 0 if there's no previous recovery
        previous_balance = last_recovery.balance if last_recovery else Decimal('0.00')

        # If there's no previous balance, set balance as total_amount - new_received_amount
        if previous_balance == Decimal('0.00'):
            balance = total_amount - new_amount_received
        else:
            # Calculate the change in total amount since the last recovery
            change_in_total_amount = total_amount - last_recovery.total_amount
    
            if change_in_total_amount == Decimal('0.00'):
            # If there's no change in total amount, use the previous balance directly
                balance = previous_balance - new_amount_received
            else:
             # Adjust the balance by adding the change in total amount to the previous balance and subtracting the new amount received
                balance = previous_balance + change_in_total_amount - new_amount_received

        # Create new recovery
        Recovery.objects.create(
            invoice=invoices.first(),  # Assuming you're working with the first invoice for simplicity
            customer=customer_name,
            total_amount=total_amount,
            amount_received=new_amount_received,
            balance=balance,
            previous_balance=previous_balance
        )
    
        # Display a message indicating that the total amount has been paid successfully
        thank = True
        return redirect('invoice:recovery-list')

    return render(request, 'invoice/recovery_form.html', {'thank': thank, 'customer_name': customer_name})


def recovery_list(request):
    # Retrieve all recoveries
    recoveries = Recovery.objects.all()
    # Create a list to store tuples of (recovery, invoice_id)
    recovery_data = [(recovery, recovery.invoice_id) for recovery in recoveries]

    return render(request, 'invoice/Recovery_list.html', {'recovery_data': recovery_data})

def delete_recovery(request, pk):
    recovery = get_object_or_404(Recovery, pk=pk)
    if request.method == 'POST':
        recovery.delete()
        return redirect('invoice:recovery-list')
    return render(request, 'invoice/delete_recovery.html', {'recovery': recovery})


# def invoice_list_recovery(request):
#     invoices = Invoice.objects.all()
#     return render(request, 'invoice/invoice_list_recovery.html', {'invoices': invoices})
def invoice_list_recovery(request):
    customer_totals = (
        Invoice.objects.values('customer')
        .annotate(total_amount=Sum('total_amount'))
        .order_by('customer')
    )
    # context={
    #     'customer_totals':customer_totals
    # }
    # print(context)

    return render(request, 'invoice/invoice_list_recovery.html', {'customer_totals': customer_totals})





# def generate_invoice_excel_report(request):
#     if request.method == 'GET':
#             # Handle GET request (render a form, for example)
#             return render(request, 'invoice/main_invoice_excel.html', context={})
#     elif request.method == 'POST':
#         # Fetch all invoices from the database
#         invoices = Invoice.objects.all()
#         # Create a new Excel workbook
#         wb = Workbook()
#         # Add a worksheet for the report
#         ws = wb.active
#         ws.title = "Invoice Report"
#         # Write headers
#         headers = [
#             "Customer", "Customer Email", "Billing Address", "Date", "Due Date", "Message",
#             "Total Amount", "Salesperson", "Manager", "Routing", "Paid Amount", "Balance",
#             "Previous Balance", "Sale Types", "Status"
#         ]
#         ws.append(headers)
#         # Write invoice data rows
#         for invoice in invoices:
#             row = [
#                 invoice.customer, invoice.customer_email, invoice.billing_address,
#                 invoice.date, invoice.due_date, invoice.message,
#                 invoice.total_amount, invoice.salesperson, invoice.manager,
#                 invoice.routing, invoice.paid_amount, invoice.balance,
#                 invoice.previous_balance, invoice.sale_types, invoice.status
#             ]
#             ws.append(row)
#         # Create HTTP response with Excel content type
#         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename=invoice_report.xlsx'
#         # Save the workbook to the HTTP response
#         wb.save(response)

#         return response
    
def generate_invoice_excel_report(request):
    if request.method == 'GET':
        # Handle GET request (render a form, for example)
        return render(request, 'invoice/main_invoice_excel.html', context={})
    elif request.method == 'POST':
        report_type = request.POST.get('report_type')
        
        if report_type == 'all':
            return generate_report_all_data(request)
        elif report_type == 'by_manager':
            manager_name = request.POST.get('manager_name')
            return generate_report_by_manager(request, manager_name)
        elif report_type == 'by_salesman':
            salesman_name = request.POST.get('salesman_name')
            return generate_report_by_salesman(request, salesman_name)
        elif report_type == 'between_dates':
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            return generate_report_between_dates(request, start_date, end_date)
        else:
            # Default behavior if no report type specified
            return HttpResponse("Invalid report type")
        
def generate_report_all_data(request):
    # Fetch all invoices from the database
    invoices = Transaction.objects.all()
    return generate_excel_report(invoices)


def generate_report_by_manager(request, manager_name):
    # Fetch invoices filtered by manager from the database
    invoices = Transaction.objects.filter(manager=manager_name)
    return generate_excel_report(invoices)

def generate_report_by_salesman(request, salesman_name):
    # Fetch invoices filtered by salesman from the database
    invoices = Transaction.objects.filter(salesperson=salesman_name)
    return generate_excel_report(invoices)

def generate_report_between_dates(request, start_date, end_date):
    # Fetch invoices between the specified dates from the database
    invoices = Transaction.objects.filter(due_date__range=[start_date, end_date])
    return generate_excel_report(invoices)

def generate_excel_report(invoices):
    # Create a new Excel workbook
    wb = Workbook()
    # Add a worksheet for the report
    ws = wb.active
    ws.title = "Invoice Report"
    # Write headers
    headers = [
        "Customer", "manager", "salesperson", "Due Date","Routing","Sale Types","Total Sales(Rs)","Payments","Balance"
    ]
    ws.append(headers)
    # Write invoice data rows
    for invoice in invoices:
        row = [
            invoice.customer, invoice.manager, invoice.salesperson,invoice.due_date, invoice.routing,invoice.sale_types, invoice.total_amount,invoice.amount,invoice.balance
        ]
        ws.append(row)
    # Create HTTP response with Excel content type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=invoice_report.xlsx'
    # Save the workbook to the HTTP response
    wb.save(response)

    return response



    
# name,address,owner_number,owner_cnic
def generate_shop_excel_report(request):
    if request.method == 'GET':
            # Handle GET request (render a form, for example)
            return render(request, 'invoice/main_shop_excel.html', context={})
    elif request.method == 'POST':
        # Fetch all invoices from the database
        shops = Shop.objects.all()
        # Create a new Excel workbook
        wb = Workbook()
        # Add a worksheet for the report
        ws = wb.active
        ws.title = "Shop List Report "
        # Write headers
        headers = [
            "name", "address", "owner_number", "owner_cnic"
        ]
        ws.append(headers)
        # Write invoice data rows
        for shop in shops:
            row = [
                shop.name, shop.address, shop.owner_number,
                shop.owner_cnic
            ]
            ws.append(row)
        # Create HTTP response with Excel content type
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=shops_report.xlsx'
        # Save the workbook to the HTTP response
        wb.save(response)

        return response
    # name,address,owner_number,owner_cnic
# def generate_recoveries_excel_report(request):
#     if request.method == 'GET':
#             # Handle GET request (render a form, for example)
#             return render(request, 'invoice/main_recoveries_excel.html', context={})
#     elif request.method == 'POST':
#         # Fetch all invoices from the database
#         recoveries = Recovery.objects.all()
#         # Create a new Excel workbook
#         wb = Workbook()
#         # Add a worksheet for the report
#         ws = wb.active
#         ws.title = "Shop List Report "
#         # Write headers
#         headers = [
#             "Customer/Shop Name", "Total Sale Amount", "Date",'Recovered Amount', "Balance",'previous Balance','approved_status'
#         ]
#         ws.append(headers)
#         # Write invoice data rowsa
#         approvedbymanager=""
#         for recovery in recoveries:
#             if recovery.approved_by_manager==False:
#                 approvedbymanager="Pending"
#             else:
#                 approvedbymanager="Approved"
            
#             row = [
#                 recovery.customer, recovery.total_amount, recovery.received_date,recovery.amount_received,
#                 recovery.balance,recovery.previous_balance,approvedbymanager
#             ]
#             ws.append(row)
#         # Create HTTP response with Excel content type
#         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename=Recoveries_report.xlsx'
#         # Save the workbook to the HTTP response
#         wb.save(response)

#         return response
    
def generate_recoveries_excel_report(request):
    if request.method == 'GET':
        # Handle GET request (render a form, for example)
        return render(request, 'invoice/main_recoveries_excel.html', context={})
    elif request.method == 'POST':
        recovery_type = request.POST.get('recovery_type')
        
        if recovery_type == 'all':
            return generate_recovery_all_data(request)
        elif recovery_type == 'by_customer':
            customer_name = request.POST.get('customer_name')
            return generate_recovery_by_customer(request, customer_name)
        elif recovery_type == 'between_dates':
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            return generate_recovery_between_dates(request, start_date, end_date)
        else:
            # Default behavior if no report type specified
            return HttpResponse("Invalid report type")
        
def generate_recovery_all_data(request):
    # Fetch all recoveries from the database
    recovery = Recovery.objects.all()
    return generate_excel_recovery(recovery)

def generate_recovery_by_customer(request, customer_name):
    # Fetch invoices filtered by manager from the database
    recovery = Recovery.objects.filter(customer=customer_name)
    return generate_excel_recovery(recovery)

def generate_recovery_between_dates(request, start_date, end_date):
    # Fetch invoices between the specified dates from the database
    recovery = Recovery.objects.filter(received_date__range=[start_date, end_date])
    return generate_excel_recovery(recovery)

def generate_excel_recovery(recovery):
    # Create a new Excel workbook
    wb = Workbook()
    # Add a worksheet for the report
    ws = wb.active
    ws.title = "Recovery Report"
    # Write headers
    headers = [
            "Customer/Shop Name", "Total Sale Amount", "Date",'Recovered Amount', "Balance",'previous Balance','approved_status'
        ]
    ws.append(headers)
    # Write invoice data rows
    for recover in recovery:
        if recover.approved_by_manager==False:
            approvedbymanager="Pending"
        else:
            approvedbymanager="Approved"
            
        row = [
            recover.customer, recover.total_amount, recover.received_date,recover.amount_received,
            recover.balance,recover.previous_balance,approvedbymanager
        ]
        ws.append(row)
    # Create HTTP response with Excel content type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=recovery_report.xlsx'
    # Save the workbook to the HTTP response
    wb.save(response)

    return response